from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from amz_researcher.models import (
    CategorySummary,
    IngredientRanking,
    ProductDetail,
    SearchProduct,
    WeightedProduct,
)

# Styling constants
HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
ACCENT_FILL = PatternFill("solid", fgColor="F5F7FA")
BORDER_BOTTOM = Border(bottom=Side(style="thin", color="D0D5DD"))
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1B2A4A")
SUBTITLE_FONT = Font(name="Arial", size=10, color="666666")
DATA_FONT = Font(name="Arial", size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
DEFAULT_ALIGN = Alignment(vertical="center")

# Item 1: Consolidated TAB_COLORS
TAB_COLORS = {
    "Market Insight": "E91E63",
    "Consumer Voice": "FF9800",
    "Badge Analysis": "673AB7",
    "Sales & Pricing": "009688",
    "Brand Positioning": "3F51B5",
    "Marketing Keywords": "795548",
    "Ingredient Ranking": "1B2A4A",
    "Category Summary": "2E86AB",
    "Rising Products": "00BCD4",
    "Product Detail": "4CAF50",
    "Raw - Search Results": "FF6B35",
    "Raw - Product Detail": "9B59B6",
}


def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = DEFAULT_ALIGN


def _style_data_rows(ws, start_row: int, end_row: int, col_count: int):
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.border = BORDER_BOTTOM
            cell.alignment = DEFAULT_ALIGN
            if (r - start_row) % 2 == 1:
                cell.fill = ACCENT_FILL


def _write_title(ws, title: str, subtitle: str, col_count: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        sub_cell = ws.cell(row=2, column=1, value=subtitle)
        sub_cell.font = SUBTITLE_FONT


def _set_column_widths(ws, widths: dict[str, float]):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _dict_to_text(d: dict) -> str:
    """{"Hair Type": "All", ...} → "Hair Type: All\nItem Form: Oil" """
    return "\n".join(f"{k}: {v}" for k, v in d.items() if not isinstance(v, (list, dict)))


def _build_ingredient_ranking(
    wb: Workbook, keyword: str,
    rankings: list[IngredientRanking],
    product_count: int,
):
    ws = wb.active
    ws.title = "Ingredient Ranking"
    ws.sheet_properties.tabColor = TAB_COLORS["Ingredient Ranking"]

    col_count = 9
    title = f"{keyword.title()} Ingredient Analysis — Weighted by Market Performance"
    subtitle = (
        f"Weight = BoughtPastMonth(30%) + BSR(25%) + Reviews(20%) + Position(15%) + Rating(10%) "
        f"| {product_count} products, {len(rankings)} ingredients"
    )
    _write_title(ws, title, subtitle, col_count)

    desc = (
        "Weighted Score: 해당 성분을 포함하는 모든 제품의 Composite Weight 합산. "
        "높을수록 시장 성과가 좋은 제품에 많이 사용됨. "
        "Avg Weight: 제품당 평균 가중치. # Products가 적어도 Avg Weight가 높으면 Top 제품에 집중된 성분."
    )
    desc_cell = ws.cell(row=3, column=1, value=desc)
    desc_cell.font = Font(name="Arial", size=9, italic=True, color="666666")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_count)

    headers = [
        "Rank", "Ingredient", "Weighted Score", "# Products",
        "Avg Weight", "Category", "Avg Price", "Price Range", "Key Insight",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=5, column=c, value=h)
    _style_header_row(ws, 5, col_count)

    for i, r in enumerate(rankings):
        row = 6 + i
        ws.cell(row=row, column=1, value=r.rank)
        ws.cell(row=row, column=2, value=r.ingredient)
        ws.cell(row=row, column=3, value=r.weighted_score).number_format = "0.000"
        ws.cell(row=row, column=4, value=r.product_count)
        ws.cell(row=row, column=5, value=r.avg_weight).number_format = "0.000"
        ws.cell(row=row, column=6, value=r.category)
        if r.avg_price is not None:
            ws.cell(row=row, column=7, value=r.avg_price).number_format = "$#,##0.00"
        ws.cell(row=row, column=8, value=r.price_range)
        ws.cell(row=row, column=9, value=r.key_insight)

    end_row = 5 + len(rankings)
    _style_data_rows(ws, 6, end_row, col_count)
    ws.freeze_panes = "A6"
    _set_column_widths(ws, {
        "A": 7, "B": 28, "C": 15, "D": 12, "E": 13,
        "F": 20, "G": 12, "H": 18, "I": 42,
    })


def _build_category_summary(wb: Workbook, categories: list[CategorySummary]):
    ws = wb.create_sheet("Category Summary")
    ws.sheet_properties.tabColor = TAB_COLORS["Category Summary"]

    col_count = 7
    _write_title(
        ws,
        "Ingredient Category Summary",
        "성분을 기능별 카테고리(Natural Oil, Vitamin, Botanical 등)로 그룹핑한 요약",
        col_count,
    )

    desc = (
        "Total Weighted Score: 카테고리 내 모든 성분의 가중치 합산 — 시장에서 해당 카테고리의 영향력. "
        "# Types: 고유 성분 수. # Mentions: 전체 제품에서 등장한 횟수 합계."
    )
    desc_cell = ws.cell(row=3, column=1, value=desc)
    desc_cell.font = Font(name="Arial", size=9, italic=True, color="666666")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_count)

    headers = [
        "Category", "Total Weighted Score", "# Types",
        "# Mentions", "Avg Price", "Price Range", "Top Ingredients",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, 4, col_count)

    for i, cat in enumerate(categories):
        row = 5 + i
        ws.cell(row=row, column=1, value=cat.category)
        ws.cell(row=row, column=2, value=cat.total_weighted_score).number_format = "0.000"
        ws.cell(row=row, column=3, value=cat.type_count)
        ws.cell(row=row, column=4, value=cat.mention_count)
        if cat.avg_price is not None:
            ws.cell(row=row, column=5, value=cat.avg_price).number_format = "$#,##0.00"
        ws.cell(row=row, column=6, value=cat.price_range)
        ws.cell(row=row, column=7, value=cat.top_ingredients)

    end_row = 4 + len(categories)
    _style_data_rows(ws, 5, end_row, col_count)
    ws.freeze_panes = "A5"
    _set_column_widths(ws, {
        "A": 22, "B": 20, "C": 10, "D": 12, "E": 12, "F": 18, "G": 45,
    })


def _build_product_detail(wb: Workbook, products: list[WeightedProduct]):
    ws = wb.create_sheet("Product Detail")
    ws.sheet_properties.tabColor = TAB_COLORS["Product Detail"]

    col_count = 20
    _write_title(
        ws,
        "Product-Level Data with Weight Breakdown",
        "각 제품의 시장 성과 지표와 Gemini가 추출한 핵심 성분 목록",
        col_count,
    )

    desc = (
        "Composite Weight: BoughtPastMonth(30%)+BSR(25%)+Reviews(20%)+Position(15%)+Rating(10%) 종합 점수. "
        "BSR(Category): Amazon 전체 카테고리 베스트셀러 순위 (낮을수록 잘 팔림)."
    )
    desc_cell = ws.cell(row=3, column=1, value=desc)
    desc_cell.font = Font(name="Arial", size=9, italic=True, color="666666")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_count)

    headers = [
        "ASIN", "Brand", "Title", "Price", "SNS Price",
        "Bought/Mo", "Reviews", "Rating", "BSR",
        "Weight", "Unit Price", "Sellers", "Coupon",
        "A+", "Badge", "Discount%", "Variations",
        "Customer Says", "Ingredients Found", "URL",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, 4, col_count)

    for i, p in enumerate(products):
        row = 5 + i
        ingredients_str = ", ".join(ing.name for ing in p.ingredients)
        ws.cell(row=row, column=1, value=p.asin)
        ws.cell(row=row, column=2, value=p.brand)
        ws.cell(row=row, column=3, value=p.title)
        if p.price is not None:
            ws.cell(row=row, column=4, value=p.price).number_format = "$#,##0.00"
        if p.sns_price is not None:
            ws.cell(row=row, column=5, value=p.sns_price).number_format = "$#,##0.00"
        if p.bought_past_month is not None:
            ws.cell(row=row, column=6, value=p.bought_past_month).number_format = "#,##0"
        ws.cell(row=row, column=7, value=p.reviews).number_format = "#,##0"
        ws.cell(row=row, column=8, value=p.rating)
        if p.bsr_category is not None:
            ws.cell(row=row, column=9, value=p.bsr_category).number_format = "#,##0"
        ws.cell(row=row, column=10, value=p.composite_weight).number_format = "0.000"
        ws.cell(row=row, column=11, value=p.unit_price)
        ws.cell(row=row, column=12, value=p.number_of_sellers)
        ws.cell(row=row, column=13, value=p.coupon)
        ws.cell(row=row, column=14, value="Y" if p.plus_content else "")
        ws.cell(row=row, column=15, value=p.badge)
        # Discount%
        if p.initial_price is not None and p.price is not None and p.initial_price > 0:
            discount_pct = round((1 - p.price / p.initial_price) * 100, 1)
            ws.cell(row=row, column=16, value=discount_pct).number_format = "0.0"
        ws.cell(row=row, column=17, value=p.variations_count)
        cs_cell = ws.cell(row=row, column=18, value=p.customer_says)
        cs_cell.alignment = WRAP_ALIGN
        ws.cell(row=row, column=19, value=ingredients_str)
        # Item 2: Product Detail URL bug fix
        url = f"https://www.amazon.com/dp/{p.asin}"
        ws.cell(row=row, column=20, value=url)

    end_row = 4 + len(products)
    _style_data_rows(ws, 5, end_row, col_count)
    ws.freeze_panes = "A5"
    _set_column_widths(ws, {
        "A": 14, "B": 16, "C": 45, "D": 10, "E": 10,
        "F": 12, "G": 10, "H": 8, "I": 10,
        "J": 10, "K": 16, "L": 8, "M": 14,
        "N": 5, "O": 18, "P": 10, "Q": 10,
        "R": 40, "S": 45, "T": 14,
    })


def _build_raw_search(wb: Workbook, keyword: str, products: list[SearchProduct]):
    ws = wb.create_sheet("Raw - Search Results")
    ws.sheet_properties.tabColor = TAB_COLORS["Raw - Search Results"]

    col_count = 8
    title = f'Amazon Search Results — "{keyword}" (Raw Data, {len(products)} products)'
    _write_title(
        ws, title,
        "Amazon 검색 결과 원본 데이터. Position은 검색 페이지 노출 순서.",
        col_count,
    )

    headers = [
        "Position", "Title", "ASIN", "Price",
        "Reviews", "Rating", "Sponsored", "Product Link",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, col_count)

    for i, p in enumerate(products):
        row = 4 + i
        ws.cell(row=row, column=1, value=p.position)
        ws.cell(row=row, column=2, value=p.title)
        ws.cell(row=row, column=3, value=p.asin)
        ws.cell(row=row, column=4, value=p.price_raw)
        ws.cell(row=row, column=5, value=p.reviews_raw)
        ws.cell(row=row, column=6, value=p.rating)
        ws.cell(row=row, column=7, value="Yes" if p.sponsored else "")
        ws.cell(row=row, column=8, value=p.product_link)

    end_row = 3 + len(products)
    _style_data_rows(ws, 4, end_row, col_count)
    ws.freeze_panes = "A4"
    _set_column_widths(ws, {
        "A": 10, "B": 60, "C": 14, "D": 12, "E": 12, "F": 8, "G": 12, "H": 50,
    })


def _build_raw_detail(wb: Workbook, details: list[ProductDetail]):
    ws = wb.create_sheet("Raw - Product Detail")
    ws.sheet_properties.tabColor = TAB_COLORS["Raw - Product Detail"]

    col_count = 10
    _write_title(
        ws,
        "Amazon Product Detail — Parsed Data (Raw)",
        "각 ASIN의 상세 페이지에서 파싱한 원본 데이터. Ingredients(raw)는 INCI 전성분 텍스트.",
        col_count,
    )

    headers = [
        "ASIN", "Brand", "BSR Category", "BSR Subcategory",
        "Rating", "Reviews", "Ingredients (raw)",
        "Features", "Measurements", "Additional Details",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, col_count)

    for i, d in enumerate(details):
        row = 4 + i
        ws.cell(row=row, column=1, value=d.asin)
        ws.cell(row=row, column=2, value=d.brand)
        if d.bsr_category is not None:
            bsr_cat_str = f"#{d.bsr_category} in {d.bsr_category_name}"
            ws.cell(row=row, column=3, value=bsr_cat_str)
        if d.bsr_subcategory is not None:
            bsr_sub_str = f"#{d.bsr_subcategory} in {d.bsr_subcategory_name}"
            ws.cell(row=row, column=4, value=bsr_sub_str)
        if d.rating is not None:
            ws.cell(row=row, column=5, value=d.rating)
        if d.review_count is not None:
            ws.cell(row=row, column=6, value=d.review_count).number_format = "#,##0"

        ing_cell = ws.cell(row=row, column=7, value=d.ingredients_raw)
        ing_cell.alignment = WRAP_ALIGN
        feat_cell = ws.cell(row=row, column=8, value=_dict_to_text(d.features))
        feat_cell.alignment = WRAP_ALIGN
        meas_cell = ws.cell(row=row, column=9, value=_dict_to_text(d.measurements))
        meas_cell.alignment = WRAP_ALIGN
        add_cell = ws.cell(row=row, column=10, value=_dict_to_text(d.additional_details))
        add_cell.alignment = WRAP_ALIGN


    end_row = 3 + len(details)
    _style_data_rows(ws, 4, end_row, col_count)
    ws.freeze_panes = "A4"
    _set_column_widths(ws, {
        "A": 14, "B": 16, "C": 30, "D": 30, "E": 8, "F": 10,
        "G": 50, "H": 35, "I": 30, "J": 35,
    })


def _build_rising_products(wb: Workbook, rising: list[dict]):
    ws = wb.create_sheet("Rising Products")
    ws.sheet_properties.tabColor = TAB_COLORS["Rising Products"]

    col_count = 8
    _write_title(
        ws,
        "Rising Products — Low Reviews, High BSR",
        "리뷰 수가 중앙값 미만이면서 BSR 10,000 이내인 제품. 신규 진입자 또는 급성장 중인 제품 후보.",
        col_count,
    )

    headers = [
        "BSR", "Brand", "Title", "Price", "Reviews",
        "Rating", "Top Ingredients", "ASIN",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, 4, col_count)

    for i, p in enumerate(rising):
        row = 5 + i
        ws.cell(row=row, column=1, value=p["bsr"]).number_format = "#,##0"
        ws.cell(row=row, column=2, value=p["brand"])
        ws.cell(row=row, column=3, value=p["title"])
        if p["price"] is not None:
            ws.cell(row=row, column=4, value=p["price"]).number_format = "$#,##0.00"
        ws.cell(row=row, column=5, value=p["reviews"]).number_format = "#,##0"
        ws.cell(row=row, column=6, value=p["rating"])
        ws.cell(row=row, column=7, value=p["top_ingredients"])
        ws.cell(row=row, column=8, value=p["asin"])

    end_row = 4 + len(rising)
    _style_data_rows(ws, 5, end_row, col_count)
    ws.freeze_panes = "A5"
    _set_column_widths(ws, {
        "A": 10, "B": 18, "C": 50, "D": 10, "E": 10,
        "F": 8, "G": 40, "H": 14,
    })


def _build_market_insight(wb: Workbook, keyword: str, report_md: str):
    """AI 시장 분석 리포트를 Market Insight 시트에 기록.

    A5 셀 하나에 전체 리포트를 넣어 Notion 복사-붙여넣기 지원.
    """
    ws = wb.create_sheet("Market Insight")
    ws.sheet_properties.tabColor = TAB_COLORS["Market Insight"]

    col_count = 1
    _write_title(
        ws,
        f"{keyword.title()} Market Insight — AI Analysis Report",
        "Powered by Gemini | Data-driven market insights",
        col_count,
    )

    note = (
        "Notion 붙여넣기: A5 셀 선택 후 상단 수식입력줄의 텍스트를 전체 선택(Ctrl+A) → 복사(Ctrl+C) → "
        "Notion에 붙여넣기하면 마크다운으로 자동 렌더링됩니다."
    )
    note_cell = ws.cell(row=3, column=1, value=note)
    note_cell.font = Font(name="Arial", size=9, italic=True, color="E91E63")

    cell = ws.cell(row=5, column=1, value=report_md)
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="top")

    line_count = report_md.count("\n") + 1
    ws.row_dimensions[5].height = min(line_count * 15, 8000)
    ws.column_dimensions["A"].width = 120
    ws.freeze_panes = "A5"


# Item 3: Consumer Voice with BSR section
def _build_consumer_voice(wb: Workbook, customer_voice_data: dict, is_keyword: bool = False):
    """customer_says 키워드 분석 결과 시트."""
    ws = wb.create_sheet("Consumer Voice")
    # Item 1: Use TAB_COLORS instead of hardcoded value
    ws.sheet_properties.tabColor = TAB_COLORS["Consumer Voice"]

    col_count = 5  # Expanded for BSR section (5 columns)
    _write_title(
        ws,
        "Consumer Voice Analysis — Keyword Sentiment",
        "Amazon AI review summary(customer_says) 기반 키워드 빈도 및 BSR 상관 분석",
        col_count,
    )

    headers = ["Keyword", "Count", "Avg BSR", "Avg Rating"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, 4, 4)

    row = 5
    # Positive keywords
    ws.cell(row=row, column=1, value="--- POSITIVE ---")
    ws.cell(row=row, column=1).font = Font(bold=True, color="2E7D32")
    row += 1
    for kw, stats in (customer_voice_data.get("positive_keywords") or {}).items():
        ws.cell(row=row, column=1, value=kw)
        ws.cell(row=row, column=2, value=stats["count"])
        if stats["avg_bsr"] is not None:
            ws.cell(row=row, column=3, value=stats["avg_bsr"]).number_format = "#,##0"
        if stats["avg_rating"] is not None:
            ws.cell(row=row, column=4, value=stats["avg_rating"])
        row += 1

    # Negative keywords
    ws.cell(row=row, column=1, value="--- NEGATIVE ---")
    ws.cell(row=row, column=1).font = Font(bold=True, color="C62828")
    row += 1
    for kw, stats in (customer_voice_data.get("negative_keywords") or {}).items():
        ws.cell(row=row, column=1, value=kw)
        ws.cell(row=row, column=2, value=stats["count"])
        if stats["avg_bsr"] is not None:
            ws.cell(row=row, column=3, value=stats["avg_bsr"]).number_format = "#,##0"
        if stats["avg_rating"] is not None:
            ws.cell(row=row, column=4, value=stats["avg_rating"])
        row += 1

    data_end_row = row - 1
    _style_data_rows(ws, 5, data_end_row, 4)

    # BSR Correlation section — 키워드 검색에서는 스킵 (크로스 카테고리 비교 무의미)
    if is_keyword:
        ws.freeze_panes = "A5"
        _set_column_widths(ws, {"A": 20, "B": 10, "C": 16, "D": 18, "E": 12})
        return

    # Item 3: BSR Correlation section
    bsr_top_pos = customer_voice_data.get("bsr_top_half_positive")
    bsr_top_neg = customer_voice_data.get("bsr_top_half_negative")
    bsr_bot_pos = customer_voice_data.get("bsr_bottom_half_positive") or {}
    bsr_bot_neg = customer_voice_data.get("bsr_bottom_half_negative") or {}

    if bsr_top_pos or bsr_top_neg:
        row += 2  # 2 blank rows
        ws.cell(row=row, column=1, value="BSR Correlation: Top Half vs Bottom Half")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        bsr_headers = ["Keyword", "Type", "Top Half Count", "Bottom Half Count", "Difference"]
        for c, h in enumerate(bsr_headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, 5)
        row += 1
        bsr_data_start = row

        # Positive keywords
        for kw, top_count in (bsr_top_pos or {}).items():
            bot_count = bsr_bot_pos.get(kw, 0)
            if top_count == 0 and bot_count == 0:
                continue
            ws.cell(row=row, column=1, value=kw)
            ws.cell(row=row, column=2, value="Positive")
            ws.cell(row=row, column=3, value=top_count)
            ws.cell(row=row, column=4, value=bot_count)
            ws.cell(row=row, column=5, value=top_count - bot_count)
            row += 1

        # Negative keywords
        for kw, top_count in (bsr_top_neg or {}).items():
            bot_count = bsr_bot_neg.get(kw, 0)
            if top_count == 0 and bot_count == 0:
                continue
            ws.cell(row=row, column=1, value=kw)
            ws.cell(row=row, column=2, value="Negative")
            ws.cell(row=row, column=3, value=top_count)
            ws.cell(row=row, column=4, value=bot_count)
            ws.cell(row=row, column=5, value=top_count - bot_count)
            row += 1

        if row > bsr_data_start:
            _style_data_rows(ws, bsr_data_start, row - 1, 5)

    ws.freeze_panes = "A5"
    _set_column_widths(ws, {"A": 20, "B": 10, "C": 16, "D": 18, "E": 12})


# Item 4: Badge Analysis with stat test and threshold sections
def _build_badge_analysis(wb: Workbook, badge_data: dict):
    """badge 보유/미보유 비교 시트."""
    ws = wb.create_sheet("Badge Analysis")
    # Item 1: Use TAB_COLORS instead of hardcoded value
    ws.sheet_properties.tabColor = TAB_COLORS["Badge Analysis"]

    col_count = 5
    _write_title(
        ws,
        "Badge Analysis — Amazon's Choice / Best Seller Impact",
        "Badge 보유 여부에 따른 BSR, 가격, 리뷰, 평점 비교",
        col_count,
    )

    headers = ["Group", "Count", "Avg BSR", "Avg Price", "Avg Rating"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, 4, col_count)

    for i, (label, key) in enumerate([("With Badge", "with_badge"), ("Without Badge", "without_badge")]):
        row = 5 + i
        metrics = badge_data.get(key, {})
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=metrics.get("count", 0))
        if metrics.get("avg_bsr") is not None:
            ws.cell(row=row, column=3, value=metrics["avg_bsr"]).number_format = "#,##0"
        if metrics.get("avg_price") is not None:
            ws.cell(row=row, column=4, value=metrics["avg_price"]).number_format = "$#,##0.00"
        if metrics.get("avg_rating") is not None:
            ws.cell(row=row, column=5, value=metrics["avg_rating"])

    # Badge types section
    row = 8
    ws.cell(row=row, column=1, value="Badge Type Distribution")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for bt in badge_data.get("badge_types", []):
        ws.cell(row=row, column=1, value=bt["badge"])
        ws.cell(row=row, column=2, value=bt["count"])
        row += 1

    _style_data_rows(ws, 5, row - 1, col_count)

    # Item 4 Section A: Statistical Test: Badge vs No-Badge BSR
    stat_test = badge_data.get("stat_test_bsr")
    if stat_test is not None:
        row += 2  # 2 blank rows
        ws.cell(row=row, column=1, value="Statistical Test: Badge vs No-Badge BSR")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        # Header row
        ws.cell(row=row, column=1, value="Test")
        ws.cell(row=row, column=2, value="Value")
        _style_header_row(ws, row, 2)
        row += 1
        stat_data_start = row

        # Method
        ws.cell(row=row, column=1, value="Method")
        ws.cell(row=row, column=2, value="Mann-Whitney U Test")
        row += 1

        # U Statistic
        ws.cell(row=row, column=1, value="U Statistic")
        if stat_test.get("u_statistic") is not None:
            ws.cell(row=row, column=2, value=stat_test["u_statistic"])
        row += 1

        # p-value
        note = stat_test.get("note", "")
        ws.cell(row=row, column=1, value="p-value")
        if note in ("insufficient_sample", "test_failed"):
            ws.cell(row=row, column=2, value=note)
        elif stat_test.get("p_value") is not None:
            ws.cell(row=row, column=2, value=stat_test["p_value"]).number_format = "0.0000"
        row += 1

        # Significant
        ws.cell(row=row, column=1, value="Significant (p < 0.05)")
        if note in ("insufficient_sample", "test_failed"):
            ws.cell(row=row, column=2, value="N/A")
        elif stat_test.get("significant") is not None:
            ws.cell(row=row, column=2, value="Yes" if stat_test["significant"] else "No")
        else:
            ws.cell(row=row, column=2, value="N/A")
        row += 1

        _style_data_rows(ws, stat_data_start, row - 1, 2)

    # Item 4 Section B: Badge Acquisition Threshold
    threshold = badge_data.get("acquisition_threshold") or {}
    if threshold:
        row += 2  # 2 blank rows
        ws.cell(row=row, column=1, value="Badge Acquisition Threshold")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        ws.cell(row=row, column=1, value="Metric")
        ws.cell(row=row, column=2, value="Value")
        _style_header_row(ws, row, 2)
        row += 1
        threshold_data_start = row

        ws.cell(row=row, column=1, value="Minimum Reviews")
        if threshold.get("min_reviews") is not None:
            ws.cell(row=row, column=2, value=threshold["min_reviews"]).number_format = "#,##0"
        row += 1

        ws.cell(row=row, column=1, value="Median Reviews")
        if threshold.get("median_reviews") is not None:
            ws.cell(row=row, column=2, value=threshold["median_reviews"]).number_format = "#,##0"
        row += 1

        ws.cell(row=row, column=1, value="Minimum Rating")
        if threshold.get("min_rating") is not None:
            ws.cell(row=row, column=2, value=threshold["min_rating"])
        row += 1

        ws.cell(row=row, column=1, value="Median Rating")
        if threshold.get("median_rating") is not None:
            ws.cell(row=row, column=2, value=threshold["median_rating"])
        row += 1

        _style_data_rows(ws, threshold_data_start, row - 1, 2)

    ws.freeze_panes = "A5"
    _set_column_widths(ws, {"A": 25, "B": 10, "C": 12, "D": 12, "E": 10})


# Item 5: Sales & Pricing sheet
def _build_sales_pricing(wb: Workbook, analysis_data: dict) -> None:
    sales = analysis_data.get("sales_volume") or {}
    sns = analysis_data.get("sns_pricing") or {}
    lt = analysis_data.get("listing_tactics") or {}
    discount = analysis_data.get("discount_impact") or {}
    promos = analysis_data.get("promotions") or {}

    # All sources empty → no sheet
    if not any([sales, sns, lt, discount, promos]):
        return

    ws = wb.create_sheet("Sales & Pricing")
    ws.sheet_properties.tabColor = TAB_COLORS["Sales & Pricing"]

    col_count = 6
    _write_title(
        ws,
        "Sales & Pricing — Revenue, Discounts & Promotions",
        "판매량, 리스팅 전술, 할인, 쿠폰 분석 통합 뷰",
        col_count,
    )

    row = 4

    # Section A: Top Sellers
    top_sellers = sales.get("top_sellers")
    if top_sellers:
        ws.cell(row=row, column=1, value="Top Sellers")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        a_headers = ["ASIN", "Brand", "Title", "Bought/Mo", "Price", "BSR"]
        for c, h in enumerate(a_headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, 6)
        row += 1
        a_data_start = row

        for ts in top_sellers:
            ws.cell(row=row, column=1, value=ts.get("asin"))
            ws.cell(row=row, column=2, value=ts.get("brand"))
            ws.cell(row=row, column=3, value=ts.get("title"))
            if ts.get("bought_past_month") is not None:
                ws.cell(row=row, column=4, value=ts["bought_past_month"]).number_format = "#,##0"
            if ts.get("price") is not None:
                ws.cell(row=row, column=5, value=ts["price"]).number_format = "$#,##0.00"
            if ts.get("bsr") is not None:
                ws.cell(row=row, column=6, value=ts["bsr"]).number_format = "#,##0"
            row += 1

        _style_data_rows(ws, a_data_start, row - 1, 6)

    # Section B: Sales by Price Tier
    price_tiers = sales.get("sales_by_price_tier")
    if price_tiers:
        row += 2
        ws.cell(row=row, column=1, value="Sales by Price Tier")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        b_headers = ["Price Tier", "Count", "Total Sales", "Avg Sales"]
        for c, h in enumerate(b_headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, 4)
        row += 1
        b_data_start = row

        tier_order = ["Budget (<$10)", "Mid ($10-25)", "Premium ($25-50)", "Luxury ($50+)"]
        for tier_name in tier_order:
            tier = price_tiers.get(tier_name)
            if tier is None:
                continue
            ws.cell(row=row, column=1, value=tier_name)
            if tier.get("count") is not None:
                ws.cell(row=row, column=2, value=tier["count"]).number_format = "#,##0"
            if tier.get("total_sales") is not None:
                ws.cell(row=row, column=3, value=tier["total_sales"]).number_format = "#,##0"
            if tier.get("avg_sales") is not None:
                ws.cell(row=row, column=4, value=tier["avg_sales"]).number_format = "#,##0"
            row += 1

        if row > b_data_start:
            _style_data_rows(ws, b_data_start, row - 1, 4)

    # Section C: Listing Tactics (keyword) or SNS Pricing (BSR)
    if lt:
        row += 2
        ws.cell(row=row, column=1, value="Listing Tactics")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        ws.cell(row=row, column=1, value="Metric")
        ws.cell(row=row, column=2, value="Value")
        ws.cell(row=row, column=3, value="Detail")
        _style_header_row(ws, row, 3)
        row += 1
        c_data_start = row

        ad = lt.get("ad_pressure") or {}
        cd = lt.get("coupon_discount") or {}
        cq = lt.get("content_quality") or {}

        kv_rows = [
            ("Sponsored Ads", f"{ad.get('sponsored_pct', '')}%", f"{ad.get('sponsored_count', 0)} of {lt.get('total_products', 0)}"),
            ("Coupon Usage", f"{cd.get('coupon_pct', '')}%", f"{cd.get('coupon_count', 0)} products"),
            ("Strikethrough Price", f"{cd.get('strikethrough_pct', '')}%", f"{cd.get('strikethrough_count', 0)} products"),
            ("A+ Content", f"{cq.get('plus_content_pct', '')}%", f"{cq.get('plus_content_count', 0)} products"),
            ("Avg Reviews", cq.get("avg_reviews"), f"median {cq.get('median_reviews', 0)}"),
            ("Avg Rating", cq.get("avg_rating"), ""),
        ]
        for label, value, detail in kv_rows:
            ws.cell(row=row, column=1, value=label)
            if value is not None:
                ws.cell(row=row, column=2, value=value)
            if detail:
                ws.cell(row=row, column=3, value=detail)
            row += 1

        # Sponsored by position sub-table
        by_pos = ad.get("by_position") or {}
        if by_pos:
            row += 1
            ws.cell(row=row, column=1, value="Sponsored by Position")
            ws.cell(row=row, column=1).font = Font(bold=True, size=10)
            row += 1
            for c, h in enumerate(["Position", "Total", "Sponsored", "Ad Rate"], 1):
                ws.cell(row=row, column=c, value=h)
            _style_header_row(ws, row, 4)
            row += 1
            for pos, d in by_pos.items():
                ws.cell(row=row, column=1, value=pos)
                ws.cell(row=row, column=2, value=d.get("total"))
                ws.cell(row=row, column=3, value=d.get("sponsored"))
                ws.cell(row=row, column=4, value=f"{d.get('sponsored_pct', 0)}%")
                row += 1

        _style_data_rows(ws, c_data_start, row - 1, 4)
    elif sns:
        row += 2
        ws.cell(row=row, column=1, value="Subscribe & Save Pricing")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        ws.cell(row=row, column=1, value="Metric")
        ws.cell(row=row, column=2, value="Value")
        _style_header_row(ws, row, 2)
        row += 1
        c_data_start = row

        retention = sns.get("retention_signal") or {}

        kv_rows = [
            ("SNS Adoption Rate", f"{sns.get('sns_adoption_pct', '')}%", None),
            ("Avg SNS Discount", f"{sns.get('avg_discount_pct', '')}%", None),
            ("SNS Avg Bought/Mo", retention.get("sns_avg_bought"), "#,##0"),
            ("No-SNS Avg Bought/Mo", retention.get("no_sns_avg_bought"), "#,##0"),
            ("With SNS Count", sns.get("with_sns_count"), "#,##0"),
            ("Without SNS Count", sns.get("without_sns_count"), "#,##0"),
        ]
        for label, value, fmt in kv_rows:
            ws.cell(row=row, column=1, value=label)
            if value is not None:
                cell = ws.cell(row=row, column=2, value=value)
                if fmt:
                    cell.number_format = fmt
            row += 1

        _style_data_rows(ws, c_data_start, row - 1, 2)

    # Section D: Discount Impact
    discount_tiers = discount.get("tiers")
    if discount_tiers:
        row += 2
        ws.cell(row=row, column=1, value="Discount Impact on BSR")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        d_headers = ["Discount Tier", "Count", "Avg BSR", "Avg Bought", "Avg Price"]
        for c, h in enumerate(d_headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, 5)
        row += 1
        d_data_start = row

        tier_order = ["No Discount (0%)", "Light (1-15%)", "Medium (16-30%)", "Heavy (31%+)"]
        for tier_name in tier_order:
            tier = discount_tiers.get(tier_name)
            if tier is None:
                continue
            ws.cell(row=row, column=1, value=tier_name)
            if tier.get("count") is not None:
                ws.cell(row=row, column=2, value=tier["count"]).number_format = "#,##0"
            if tier.get("avg_bsr") is not None:
                ws.cell(row=row, column=3, value=tier["avg_bsr"]).number_format = "#,##0"
            if tier.get("avg_bought") is not None:
                ws.cell(row=row, column=4, value=tier["avg_bought"]).number_format = "#,##0"
            if tier.get("avg_price") is not None:
                ws.cell(row=row, column=5, value=tier["avg_price"]).number_format = "$#,##0.00"
            row += 1

        if row > d_data_start:
            _style_data_rows(ws, d_data_start, row - 1, 5)

    # Section E: Coupon Types
    coupon_types = promos.get("coupon_types")
    if coupon_types:
        row += 2
        ws.cell(row=row, column=1, value="Coupon Type Distribution")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        ws.cell(row=row, column=1, value="Coupon")
        ws.cell(row=row, column=2, value="Count")
        _style_header_row(ws, row, 2)
        row += 1
        e_data_start = row

        for ct in coupon_types:
            ws.cell(row=row, column=1, value=ct.get("coupon"))
            if ct.get("count") is not None:
                ws.cell(row=row, column=2, value=ct["count"]).number_format = "#,##0"
            row += 1

        if row > e_data_start:
            _style_data_rows(ws, e_data_start, row - 1, 2)

    ws.freeze_panes = "A5"
    _set_column_widths(ws, {
        "A": 22, "B": 16, "C": 40, "D": 14, "E": 12, "F": 10,
    })


# Item 6: Brand Positioning sheet
def _build_brand_positioning_sheet(wb: Workbook, analysis_data: dict) -> None:
    positioning = analysis_data.get("brand_positioning")  # list[dict]
    mfr = analysis_data.get("manufacturer") or {}

    if not positioning and not mfr:
        return

    ws = wb.create_sheet("Brand Positioning")
    ws.sheet_properties.tabColor = TAB_COLORS["Brand Positioning"]

    col_count = 7
    _write_title(
        ws,
        "Brand Positioning — Price vs BSR Analysis",
        "브랜드/제조사별 가격 포지셔닝 및 시장 성과 비교",
        col_count,
    )

    row = 4

    # Section A: Brand Positioning
    if positioning:
        ws.cell(row=row, column=1, value="Brand Positioning")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        a_headers = ["Brand", "Products", "Avg Price", "Avg BSR", "Avg Rating", "Total Reviews", "Segment"]
        for c, h in enumerate(a_headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, 7)
        row += 1
        a_data_start = row

        for bp in positioning:
            ws.cell(row=row, column=1, value=bp.get("brand"))
            if bp.get("product_count") is not None:
                ws.cell(row=row, column=2, value=bp["product_count"]).number_format = "#,##0"
            if bp.get("avg_price") is not None:
                ws.cell(row=row, column=3, value=bp["avg_price"]).number_format = "$#,##0.00"
            if bp.get("avg_bsr") is not None:
                ws.cell(row=row, column=4, value=bp["avg_bsr"]).number_format = "#,##0"
            if bp.get("avg_rating") is not None:
                ws.cell(row=row, column=5, value=bp["avg_rating"]).number_format = "0.00"
            if bp.get("total_reviews") is not None:
                ws.cell(row=row, column=6, value=bp["total_reviews"]).number_format = "#,##0"
            ws.cell(row=row, column=7, value=bp.get("segment"))
            row += 1

        _style_data_rows(ws, a_data_start, row - 1, 7)

    # Section B: Top Manufacturers
    top_mfrs = mfr.get("top_manufacturers")
    if top_mfrs:
        row += 2
        ws.cell(row=row, column=1, value="Top Manufacturers")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        b_headers = ["Manufacturer", "Products", "Avg BSR", "Avg Price", "Avg Rating", "Total Bought", "K-Beauty"]
        for c, h in enumerate(b_headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, 7)
        row += 1
        b_data_start = row

        for m in top_mfrs:
            ws.cell(row=row, column=1, value=m.get("manufacturer"))
            if m.get("product_count") is not None:
                ws.cell(row=row, column=2, value=m["product_count"]).number_format = "#,##0"
            if m.get("avg_bsr") is not None:
                ws.cell(row=row, column=3, value=m["avg_bsr"]).number_format = "#,##0"
            if m.get("avg_price") is not None:
                ws.cell(row=row, column=4, value=m["avg_price"]).number_format = "$#,##0.00"
            if m.get("avg_rating") is not None:
                ws.cell(row=row, column=5, value=m["avg_rating"]).number_format = "0.00"
            if m.get("total_bought") is not None:
                ws.cell(row=row, column=6, value=m["total_bought"]).number_format = "#,##0"
            ws.cell(row=row, column=7, value="Y" if m.get("is_kbeauty") else "")
            row += 1

        _style_data_rows(ws, b_data_start, row - 1, 7)

    # Section C: Market Concentration
    mc = mfr.get("market_concentration")
    if mc:
        row += 2
        ws.cell(row=row, column=1, value="Market Concentration")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        ws.cell(row=row, column=1, value="Metric")
        ws.cell(row=row, column=2, value="Value")
        _style_header_row(ws, row, 2)
        row += 1
        c_data_start = row

        ws.cell(row=row, column=1, value="Total Manufacturers")
        if mfr.get("total_manufacturers") is not None:
            ws.cell(row=row, column=2, value=mfr["total_manufacturers"]).number_format = "#,##0"
        row += 1

        ws.cell(row=row, column=1, value="Top 10 Products")
        if mc.get("top10_products") is not None:
            ws.cell(row=row, column=2, value=mc["top10_products"]).number_format = "#,##0"
        row += 1

        ws.cell(row=row, column=1, value="Total Products")
        if mc.get("total_products") is not None:
            ws.cell(row=row, column=2, value=mc["total_products"]).number_format = "#,##0"
        row += 1

        ws.cell(row=row, column=1, value="Top 10 Market Share")
        if mc.get("top10_share_pct") is not None:
            ws.cell(row=row, column=2, value=f"{mc['top10_share_pct']}%")
        row += 1

        _style_data_rows(ws, c_data_start, row - 1, 2)

    ws.freeze_panes = "A5"
    _set_column_widths(ws, {
        "A": 24, "B": 10, "C": 12, "D": 12, "E": 10, "F": 14, "G": 18,
    })


# Item 7: Marketing Keywords sheet
def _build_marketing_keywords(wb: Workbook, analysis_data: dict) -> None:
    kw_data = analysis_data.get("title_keywords") or {}
    tier_data = analysis_data.get("price_tier_analysis") or {}

    if not kw_data and not tier_data:
        return

    ws = wb.create_sheet("Marketing Keywords")
    ws.sheet_properties.tabColor = TAB_COLORS["Marketing Keywords"]

    col_count = 4
    _write_title(
        ws,
        "Marketing Keywords — Title Keyword Performance",
        "제품 타이틀 내 마케팅 키워드별 BSR/판매량 + 가격대별 Top 성분",
        col_count,
    )

    row = 4

    # Section A: Title Keyword Performance
    keyword_analysis = kw_data.get("keyword_analysis")
    if keyword_analysis:
        ws.cell(row=row, column=1, value="Title Keyword Performance")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        a_headers = ["Keyword", "Count", "Avg BSR", "Avg Bought/Mo"]
        for c, h in enumerate(a_headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, 4)
        row += 1
        a_data_start = row

        for kw, metrics in keyword_analysis.items():
            ws.cell(row=row, column=1, value=kw)
            if metrics.get("count") is not None:
                ws.cell(row=row, column=2, value=metrics["count"]).number_format = "#,##0"
            if metrics.get("avg_bsr") is not None:
                ws.cell(row=row, column=3, value=metrics["avg_bsr"]).number_format = "#,##0"
            if metrics.get("avg_bought") is not None:
                ws.cell(row=row, column=4, value=metrics["avg_bought"]).number_format = "#,##0"
            row += 1

        if row > a_data_start:
            _style_data_rows(ws, a_data_start, row - 1, 4)

    # Section B: Price Tier Top Ingredients
    if tier_data:
        row += 2
        ws.cell(row=row, column=1, value="Price Tier Top Ingredients")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        b_headers = ["Price Tier", "Products", "Top Ingredients"]
        for c, h in enumerate(b_headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header_row(ws, row, 3)
        row += 1
        b_data_start = row

        tier_order = ["Budget (<$10)", "Mid ($10-25)", "Premium ($25-50)", "Luxury ($50+)"]
        for tier_name in tier_order:
            tier = tier_data.get(tier_name)
            if tier is None:
                continue
            ws.cell(row=row, column=1, value=tier_name)
            if tier.get("product_count") is not None:
                ws.cell(row=row, column=2, value=tier["product_count"]).number_format = "#,##0"
            top_ings = tier.get("top_ingredients") or []
            ing_str = ", ".join(ing["name"] for ing in top_ings if isinstance(ing, dict) and "name" in ing)
            ing_cell = ws.cell(row=row, column=3, value=ing_str)
            ing_cell.alignment = WRAP_ALIGN
            row += 1

        if row > b_data_start:
            _style_data_rows(ws, b_data_start, row - 1, 3)

    ws.freeze_panes = "A5"
    _set_column_widths(ws, {
        "A": 22, "B": 10, "C": 14, "D": 14,
    })


# Item 8: Updated build_excel()
def build_excel(
    keyword: str,
    weighted_products: list[WeightedProduct],
    rankings: list[IngredientRanking],
    categories: list[CategorySummary],
    search_products: list[SearchProduct],
    details: list[ProductDetail],
    market_report: str = "",
    rising_products: list[dict] | None = None,
    # V5 추가
    analysis_data: dict | None = None,
) -> bytes:
    wb = Workbook()

    # === Insight tabs ===
    _build_ingredient_ranking(wb, keyword, rankings, len(weighted_products))
    if market_report:
        _build_market_insight(wb, keyword, market_report)
    if analysis_data:
        customer_voice = analysis_data.get("customer_voice")
        if customer_voice:
            _build_consumer_voice(wb, customer_voice)
        badge_data = analysis_data.get("badges")
        if badge_data:
            _build_badge_analysis(wb, badge_data)
        _build_sales_pricing(wb, analysis_data)

    # === Analysis tabs ===
    if analysis_data:
        _build_brand_positioning_sheet(wb, analysis_data)
        _build_marketing_keywords(wb, analysis_data)
    _build_category_summary(wb, categories)
    if rising_products:
        _build_rising_products(wb, rising_products)
    _build_product_detail(wb, weighted_products)

    # === Raw tabs ===
    _build_raw_search(wb, keyword, search_products)
    _build_raw_detail(wb, details)

    # Reorder sheets
    desired_order = [
        "Market Insight",
        "Consumer Voice",
        "Badge Analysis",
        "Sales & Pricing",
        "Brand Positioning",
        "Marketing Keywords",
        "Ingredient Ranking",
        "Category Summary",
        "Rising Products",
        "Product Detail",
        "Raw - Search Results",
        "Raw - Product Detail",
    ]
    # Filter to only existing sheets, preserve any unlisted sheets at the end
    existing = wb.sheetnames
    ordered = [s for s in desired_order if s in existing]
    ordered += [s for s in existing if s not in ordered]
    wb._sheets = [wb[s] for s in ordered]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_keyword_excel(
    keyword: str,
    weighted_products: list[WeightedProduct],
    rankings: list[IngredientRanking],
    categories: list[CategorySummary],
    search_products: list[SearchProduct],
    details: list[ProductDetail],
    market_report: str = "",
    analysis_data: dict | None = None,
) -> bytes:
    """키워드 검색 전용 9시트 Excel 생성.

    카테고리 리포트(12시트)에서 BSR 의존 3시트 제거:
    - Badge Analysis (삭제)
    - Brand Positioning (삭제)
    - Rising Products (삭제)

    Consumer Voice는 BSR correlation 섹션만 제거 (is_keyword=True).
    """
    wb = Workbook()

    # === Insight tabs ===
    _build_ingredient_ranking(wb, keyword, rankings, len(weighted_products))
    if market_report:
        _build_market_insight(wb, keyword, market_report)
    if analysis_data:
        customer_voice = analysis_data.get("customer_voice")
        if customer_voice:
            _build_consumer_voice(wb, customer_voice, is_keyword=True)
        # Badge Analysis 제거
        _build_sales_pricing(wb, analysis_data)

    # === Analysis tabs ===
    if analysis_data:
        # Brand Positioning 제거
        _build_marketing_keywords(wb, analysis_data)
    _build_category_summary(wb, categories)
    # Rising Products 제거
    _build_product_detail(wb, weighted_products)

    # === Raw tabs ===
    _build_raw_search(wb, keyword, search_products)
    _build_raw_detail(wb, details)

    # Reorder: 9시트
    desired_order = [
        "Market Insight",
        "Consumer Voice",
        "Sales & Pricing",
        "Marketing Keywords",
        "Ingredient Ranking",
        "Category Summary",
        "Product Detail",
        "Raw - Search Results",
        "Raw - Product Detail",
    ]
    existing = wb.sheetnames
    ordered = [s for s in desired_order if s in existing]
    ordered += [s for s in existing if s not in ordered]
    wb._sheets = [wb[s] for s in ordered]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
